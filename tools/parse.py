# -*- coding: utf-8 -*-
import re, json, glob, os
from openpyxl import load_workbook

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")
vocab=[]
files=sorted(glob.glob(BASE+"/Quiz/*.txt"))
for f in files:
    m=re.search(r'第(\d+)組_(.+)\.txt', os.path.basename(f))
    gid, pos = int(m.group(1)), m.group(2)
    for line in open(f, encoding='utf-8'):
        line=line.rstrip('\n').strip()
        if not line: continue
        parts=line.split('\t')
        word=parts[0].strip()
        rest=parts[1] if len(parts)>1 else ''
        # split example
        ex=''
        if '／例:' in rest:
            rest, ex = rest.split('／例:',1)
        elif '/例:' in rest:
            rest, ex = rest.split('/例:',1)
        # rest = "よみ／english 中文"  or  "english 中文" (外來語)
        reading=''
        if '／' in rest:
            reading, rest2 = rest.split('／',1)
        else:
            reading=''; rest2=rest
        rest2=rest2.strip()
        # english part = leading ascii/latin; chinese = rest
        mm=re.match(r'^([A-Za-z0-9 ,\.\-\'\(\)/]+?)\s+([^\x00-\x7F].*)$', rest2)
        if mm: en, zh = mm.group(1).strip(), mm.group(2).strip()
        else: en, zh = '', rest2
        vocab.append(dict(id=len(vocab)+1, group=gid, pos=pos, word=word,
                          reading=reading.strip() or word, en=en, zh=zh, ex=ex.strip()))
print("vocab", len(vocab))
json.dump(vocab, open(os.path.join(os.path.dirname(os.path.abspath(__file__)),'vocab_raw.json'),'w',encoding='utf-8'), ensure_ascii=False, indent=1)

wb=load_workbook(BASE+"/N2文法144_上篇.xlsx"); ws=wb.active
gram=[]
for r in range(2, ws.max_row+1):
    n=ws.cell(r,1).value; pat=ws.cell(r,2).value; mean=ws.cell(r,3).value
    use=ws.cell(r,4).value; ex=ws.cell(r,5).value
    note=''
    if '\n注意：' in mean:
        mean, note = mean.split('\n注意：',1)
    exj, exz = ex, ''
    if '\n' in ex:
        exj, exz = ex.split('\n',1)
    exz=exz.strip().strip('（）()')
    gram.append(dict(id=n, pattern=pat.strip(), meaning=mean.strip(), note=note.strip(),
                     usage=[u.strip() for u in use.split('\n') if u.strip()],
                     ex=exj.strip(), exZh=exz))
print("grammar", len(gram))
json.dump(gram, open(os.path.join(os.path.dirname(os.path.abspath(__file__)),'gram_raw.json'),'w',encoding='utf-8'), ensure_ascii=False, indent=1)

print("→ 接著執行 gen_data.py 產生 data/*.js")
